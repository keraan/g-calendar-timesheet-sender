import shutil
from datetime import date, timedelta, datetime
from pytz import timezone
from openpyxl import load_workbook
import re
import json

## Global Variables
def load_config(filename):
    with open(filename, "r") as config_file:
        return json.load(config_file)

config = load_config("config.json")
my_firstname = config["my_firstname"]
my_lastname = config["my_lastname"]
my_email = config["my_email"]
to_email = config["recipient_email"]
to_name = config["to_name"]


def input_day(ws, day, start, end):
    start_cell = ""
    end_cell = ""
    total_hours_cell = ""
    day = day.lower()
    if (day == "mon" or day == "monday"):
        start_cell = "C7"
        end_cell = "D7"
        total_hours_cell = "F7"
    elif (day == "tues" or day == "tuesday"):
        start_cell = "C8"
        end_cell = "D8"
        total_hours_cell = "F8"
    elif (day == "wed" or day == "wednesday"):
        start_cell = "C9"
        end_cell = "D9"
        total_hours_cell = "F9"
    elif (day == "thurs" or day == "thursday"):
        start_cell = "C10"
        end_cell = "D10"
        total_hours_cell = "F10"
    elif (day == "fri" or day == "friday"):
        start_cell = "C11"
        end_cell = "D11"
        total_hours_cell = "F11"
    elif (day == "sat" or day == "saturday"):
        start_cell = "C12"
        end_cell = "D12"
        total_hours_cell = "F12"
    
    if start_cell and end_cell:
        ws[start_cell].number_format = 'h:mm AM/PM'
        ws[end_cell].number_format = 'h:mm AM/PM'
        
        # Convert Excel float time to datetime time
        start_time = (datetime.min + timedelta(days=start)).time()
        end_time = (datetime.min + timedelta(days=end)).time()

        # Convert datetime time to string with AM/PM
        start_time_str = start_time.strftime("%I:%M %p")
        end_time_str = end_time.strftime("%I:%M %p")
        
        ws[start_cell] = start_time_str
        ws[end_cell] = end_time_str

        # Calculate total hours
        total_hours = end - start

        ws[total_hours_cell] = format_time_as_string(total_hours)


def format_time_as_string(hours_float):
    total_minutes = round(hours_float * 24 * 60)  # Step 1: Round to the nearest minute
    hours = total_minutes // 60
    minutes = total_minutes % 60  # Step 2: Convert back to hours and minutes
    return f"{hours}h {minutes}m"



def time_to_excel_float(time_str):
    if ":" in time_str:
        hours, minutes = map(int, time_str.split(":"))
        return hours / 24.0 + (minutes / 1440.0)  # 1440 = 24 * 60
    else:
        hours = int(time_str)
        return hours / 24.0


def parse_time_str(time_str):
    match = re.search(r'(\d+)h (\d+)m', time_str)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        total_minutes = hours * 60 + minutes
        return total_minutes
    else:
        return 0  # Return 0 if the format doesn't match


def get_total_hours(ws):
    total_minutes = 0
    for cell in ["F7", "F8", "F9", "F10", "F11", "F12"]:
        cell_value = ws[cell].value
        if cell_value is not None:
            total_minutes += parse_time_str(cell_value)
    
    total_hours = total_minutes // 60
    remaining_minutes = total_minutes % 60
    return f"{total_hours}h {remaining_minutes}m"


def convert_time_to_excel_float(time_str):
    local_time = datetime.fromisoformat(time_str)
    midnight = datetime.combine(local_time.date(), datetime.min.time()).replace(tzinfo=local_time.tzinfo)
    delta = local_time - midnight
    return delta.total_seconds() / 86400.0  # 86400 seconds in a day


def load_timesheet(days_worked_data):
    original_file = config["original_file"]
    file_name = f"{(date.today() - timedelta(days=7)).strftime('%d-%m-%Y')} to {date.today().strftime('%d-%m-%Y')} Timesheet - {my_firstname} {my_lastname}.xlsx"

    shutil.copy(original_file, file_name)
    
    wb = load_workbook(file_name)
    ws = wb.active
    
    ws["B4"] = my_firstname
    ws["C4"] = my_lastname
    ws["F4"] = date.today().strftime("%d-%m-%y")

    for day, times in days_worked_data.items():
        start_time_str, end_time_str, time_zone_str = times
        start_time_float = convert_time_to_excel_float(start_time_str)
        end_time_float = convert_time_to_excel_float(end_time_str)
        input_day(ws, day, start_time_float, end_time_float)

    ws["F14"] = get_total_hours(ws)
    wb.save(file_name)


if __name__ == "__main__":
    with open("google-calendar-info-getter.py", "r") as f:
        code = f.read()
        exec(code)
    
    work_data = {}

    try:
        with open("work_data.json", "r") as f:
            work_data = json.load(f)
    except FileNotFoundError:
        work_data = {}

    load_timesheet(work_data)